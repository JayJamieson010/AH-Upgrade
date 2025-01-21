import json
from PyQt5.QtWidgets import QWidget, QLabel, QVBoxLayout, QLineEdit, QPushButton, QFileDialog, QMessageBox
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import pyautogui
import time
import os   

# Global variables for user inputs
PATH = ""
email = ""
password = ""
testing = False

#Testing mode inclution

CONFIG_FILE = "Main\Xero.json"  # Path to the JSON file where we will save the variables

def load_config():
    """Load configuration from a JSON file."""
    global PATH, email, password
    try:
        with open(CONFIG_FILE, "r") as file:
            config = json.load(file)
            PATH = config.get("path", "")
            email = config.get("email", "")
            password = config.get("password", "")
    except FileNotFoundError:
        print("Configuration file not found. Using default values.")
    except json.JSONDecodeError:
        print("Error reading the configuration file.")

def save_config():
    """Save configuration to a JSON file, creating the file if it doesn't exist."""
    config = {
        "path": PATH,
        "email": email,
        "password": password
    }
    try:
        # Ensure the directory structure exists
        os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)

        with open(CONFIG_FILE, "w") as file:
            json.dump(config, file, indent=4)
    except Exception as e:
        print(f"Error saving configuration: {e}")

def xero_setup():
    """
    Function to create a small window for setting the PATH, email, and password variables.
    """
    setup_window = QWidget()
    setup_window.setWindowTitle("Xero Setup")
    setup_window.setMinimumSize(400, 300)

    # Apply Xero-inspired styles
    xero_blue = "#0074D9"
    setup_window.setStyleSheet(f"""
        QWidget {{
            background-color: #F4F4F9;
        }}
        QLabel {{
            font-size: 14px;
            color: #333333;
        }}
        QLineEdit {{
            border: 1px solid {xero_blue};
            border-radius: 5px;
            padding: 5px;
            font-size: 14px;
        }}
        QPushButton {{
            background-color: {xero_blue};
            color: white;
            border: none;
            border-radius: 5px;
            padding: 10px;
            font-size: 14px;
        }}
        QPushButton:hover {{
            background-color: #0057A4;
        }}
        QPushButton:pressed {{
            background-color: #003E73;
        }}
    """)

    layout = QVBoxLayout()

    # Input fields
    path_label = QLabel("Chromedriver PATH:")
    path_input = QLineEdit()
    path_input.setPlaceholderText("Enter the path to chromedriver.exe")
    path_input.setText(PATH)

    email_label = QLabel("Email:")
    email_input = QLineEdit()
    email_input.setPlaceholderText("Enter your Xero email")
    email_input.setText(email)

    password_label = QLabel("Password:")
    password_input = QLineEdit()
    password_input.setPlaceholderText("Enter your Xero password")
    password_input.setEchoMode(QLineEdit.Password)
    password_input.setText(password)

    # Function to save inputs
    def save_inputs():
        global PATH, email, password
        PATH = path_input.text().strip()
        email = email_input.text().strip()
        password = password_input.text().strip()

        if not PATH or not email or not password:
            QMessageBox.warning(setup_window, "Warning", "All fields are required!")
        else:
            save_config()  # Save configuration to JSON file
            QMessageBox.information(setup_window, "Success", "Setup completed!")
            setup_window.close()

    # Save button
    save_button = QPushButton("Save")
    save_button.clicked.connect(save_inputs)

    # Add widgets to the layout
    layout.addWidget(path_label)
    layout.addWidget(path_input)
    layout.addWidget(email_label)
    layout.addWidget(email_input)
    layout.addWidget(password_label)
    layout.addWidget(password_input)
    layout.addWidget(save_button)

    setup_window.setLayout(layout)
    setup_window.show()


# Global variable to keep the window reference
xero_window = None

def create_window():
    
    """Function to create and display the Xero Automation window."""
    global xero_window  # Use a global variable to hold the window reference

    # Create a new widget as a window
    xero_window = QWidget()
    xero_window.setWindowTitle("Xero Automation Window")
    xero_window.setMinimumSize(400, 300)

    # Variables to hold selected file path
    selected_file_path = None

    # Function to browse and select an Excel file
    def clear_and_write(content):
        try:
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.5)
            pyautogui.press('backspace')
            time.sleep(0.5)

            print(f"Writing content: {content}")
            pyautogui.write(content)
            time.sleep(1)
            print(f"Successfully wrote content: {content}")
        except Exception as e:
            print(f"Error writing content: {e}")

    def locate_element(image_path, confidence=0.8, timeout=10):
        print(f"Locating element: {image_path}")
        start_time = time.time()
        while time.time() - start_time < timeout:
            element = pyautogui.locateCenterOnScreen(image_path, confidence=confidence)
            if element:
                print(f"Element {image_path} located.")
                return element
            time.sleep(0.5)
        print(f"Failed to locate element: {image_path}")
        return None

    def browse_file():
        nonlocal selected_file_path
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(
            xero_window, "Select Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options
        )
        if file_name:
            excel_path_label.setText(f"Selected File: {file_name}")
            selected_file_path = file_name
        else:
            excel_path_label.setText("No file selected.")
            selected_file_path = None

    # Xero Automation function
    def xero_statement_sender():
        global email, password, PATH  # Declare global variables to modify them inside the function

        load_config()  # This will load the config and set the global variables

    # Check if email and password have been loaded correctly
        if email and password:
            print(f"Email: {email}, Password: {password}")
        else:
            print("Did not load configuration properly.")
        """Runs the Xero automation process."""

        time.sleep(3)
        if not selected_file_path:
            QMessageBox.warning(xero_window, "Warning", "Please select an Excel file.")
            return

        try:
            # Path to your chromedriver.exe file
            global PATH
            service = Service(PATH)
            driver = webdriver.Chrome(service=service)

            wb = load_workbook(selected_file_path)
            ws = wb.active
            student_list = list(ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True))
            end_date_selected = pyautogui.prompt("End Date?", "Text")
            testing_message = pyautogui.prompt("Are you running a test?").lower()
            if testing_message == "yes":
                global testing
                testing = True
                print (testing)
            
            else:
                print("not testing") 
                print(testing) 
            time.sleep(3)
            # Maximize the browser window
            driver.maximize_window()

            # Navigate to the Xero login page
            driver.get("https://login.xero.com/")
            time.sleep(3)

            # Login process
            email_field = driver.find_element(By.CLASS_NAME, "xui-textinput--input")
            password_field = driver.find_element(By.CLASS_NAME, "xl-form-password")
            email_field.send_keys(email)
            password_field.send_keys(password)
            login_button = driver.find_element(By.CLASS_NAME, "xui-button")
            login_button.click()
            time.sleep(2)

            driver.execute_script("window.scrollBy(0, 500);")
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "xui-margin-bottom-xsmall")))
            time.sleep(10)
            # Check for MFA requirement
            mfa_element = driver.find_elements(By.CLASS_NAME, "xui-margin-bottom-xsmall")
            if mfa_element:
                pyautogui.scroll(-200)
                use_backup = driver.find_element(By.XPATH, "//button[text()='Use a backup method instead']")  # Adjust XPath if needed
                if use_backup:
                    use_backup.click()
                    backup_email = driver.find_element(By.XPATH, "//h2[text()='Backup email address']") 
                    time.sleep(1)
                    backup_email.click()
                    send_backup_code = driver.find_element(By.XPATH, "//button[text()='Send code']") 
                    time.sleep(1)
                    send_backup_code.click()
                    time.sleep(30)
                    pyautogui.press("tab")
                    pyautogui.press("enter")

                time.sleep(20)  # Wait for page to load
                contacts_button = driver.find_element(By.XPATH, "//button[text()='Contacts']")  # Adjust XPath if needed
                if contacts_button:
                    contacts_button.click()
                    time.sleep(2)

                    all_contacts = driver.find_element(By.XPATH, "//a[text()='All contacts']") 

                    if all_contacts:
                        all_contacts.click()
                        time.sleep(3)

                        options = driver.find_element(By.CLASS_NAME, "xui-touchtarget")
                        if options:
                            options.click()
                            time.sleep(3)

                            send_statement = driver.find_element(By.XPATH, "//span[text()='Send statements']")
                            if send_statement:
                                send_statement.click()
                                time.sleep(3)
            else:
                time.sleep(20)  # Wait for page to load
                contacts_button = driver.find_element(By.XPATH, "//button[text()='Contacts']")  # Adjust XPath if needed
                # Click the button
                if contacts_button:
                    contacts_button.click()
                    time.sleep(2)  # Pause to observe or wait for the next action

                    all_contacts = driver.find_element(By.XPATH, "//a[text()='All contacts']") 

                    if all_contacts:
                        all_contacts.click()
                        time.sleep(3)

                        #options = locate_element("Main/Images/options.png", confidence=0.8)
                        options = driver.find_element(By.CLASS_NAME, "xui-touchtarget")
                        if options:
                            options.click()
                            time.sleep(3)

                            send_statement = driver.find_element(By.XPATH, "//span[text()='Send statements']")  # Adjust XPath if needed
                            if send_statement:
                                send_statement.click()
                                time.sleep(3)

            # Additional Xero steps with Excel data integration
            for student_tuple in student_list:
                student_number = student_tuple[0]
                if student_number is None:
                    continue
                print(f"Processing student number: {student_number}")
                time.sleep(4)

                activity_field = driver.find_element(By.ID, "StatementTypeFromForm_value")
                if activity_field:
                    print("Found")
                    activity_field.click()
                    # Ctrl + A and Backspace using Selenium
                    activity_field.send_keys(Keys.CONTROL + 'a')  # Select all text
                    activity_field.send_keys(Keys.BACKSPACE)  # Clear the text
                    activity_field.send_keys("Activity")  # Type "Activity

                    # Navigate through the UI and send the statement
                    # Additional logic for Excel data processing can go here
                    pyautogui.press('tab')
                    pyautogui.write("1 Jan 2023")
                    pyautogui.press('tab')
                    pyautogui.write(end_date_selected)
                    pyautogui.press('tab')
                    clear_and_write(str(student_number))
                    pyautogui.press("tab")
                    pyautogui.press("enter")

                    time.sleep(3)

                            # Assuming driver is your initialized Selenium WebDriver instance
                    checkboxes = driver.find_elements(By.ID, "ext-gen26")

                            # Loop through each checkbox and click it if it is not already selected
                    for checkbox in checkboxes:
                        if not checkbox.is_selected():
                            checkbox.click()

                            emailButton = driver.find_element(By.ID, "ext-gen22")
                            if emailButton:
                                emailButton.click()
                                time.sleep(3)
                            
                            emailAdress = driver.find_element(By.XPATH, "//*[contains(@id, 'MessageTo')]")

                            if emailAdress:
                                emailAdress.click()
                                emailAdress.send_keys(Keys.CONTROL + 'a')  # Select all text
                                emailAdress.send_keys(Keys.BACKSPACE)
                                if testing:
                                    emailAdress.send_keys(email)
                

                            sendButton= driver.find_element(By.ID, "email01")
                            if sendButton:
                                sendButton.click()
                                
                else:
                               time.sleep(20)# Wait for page to load 

                print(f"Statement sent for student number: {student_number}")

            

            # Notify user upon completion
            QMessageBox.information(xero_window, "Success", "Statements have been sent.")
        except Exception as e:
            print(f"Error: {e}")
            QMessageBox.warning(xero_window, "Error", f"An error occurred: {e}")

    # Excel browsing setup
    browse_button = QPushButton("Browse Excel File")
    browse_button.clicked.connect(browse_file)

    excel_path_label = QLabel("No file selected.")
    
    # Create the layout for the window
    layout = QVBoxLayout()
    layout.addWidget(excel_path_label)
    layout.addWidget(browse_button)
    layout.addWidget(QPushButton("Send Statement", clicked=xero_statement_sender))
    xero_window.setLayout(layout)
    xero_window.show()

# Main entry point
if __name__ == "__main__":
    load_config()
    xero_setup()  # To setup and configure the user's preferences
    create_window()  # Create the Xero automation window
